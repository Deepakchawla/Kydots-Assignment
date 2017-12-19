
class JobRecommendedSysetm:
	def __init__(self):
		import requests
		from sklearn.cross_validation import train_test_split
		import pandas as pd
		from bs4 import BeautifulSoup
		import openpyxl
		import random	
		import xlrd
		import csv
		import pandas as pd
		import graphlab
"""
This function crawl the institute names and save in the list from the given url using Python BeautifulSoup library.
"""
	def crawl_institude_names(self):
		url = 'https://www.4icu.org/in/indian-universities.htm' #Url for html content parsing.
		page = requests.get(url)
		soup = BeautifulSoup(page.text, 'html.parser') #It parses the html data from requested url.
		institude_html_data = soup.find(class_='table table-hover') #It parses html data of specific class from all web page data.
		institude_html_data = institude_html_data.find_all('a') #It parses html data of anchor tag from all html data with specific class.
		institude_names = []
		for temp_name in institude_html_data:
		    institude_names.append(temp_name.contents[0]) #It appends anchor tag content in the institude_name list.
		insitude_names = insitude_names[5:]
"""
This function crawl the degree names and save in the list from the given url using Python BeautifulSoup library.
"""
	def crawl_degree_names(self):		
		url = 'https://targetstudy.com/degree/' #Url for html content parsing.
		headers = {'User-agent': 'Mozilla/5.0'} #Indulge Headers because this website doesn't give permission to access its html data.
		page = requests.get(url, headers=headers)
		soup = BeautifulSoup(page.text, 'html.parser')  #It parses the html data from requested url.
		degree_html_data = soup.find(class_='panel-body') #It parses html data of specific class from all web page data.
		degree_html_data = degree__html_data.find_all('li') #It parses html data of list tag from all html data with specific class.
		degree_names = []
		index = 0
		for temp_name in degree_html_data:
		    try: #Try block if any TypeError error generate in for lop.
			degree_names.append(temp_name.contents[0].replace('\xa0', '')) #It replace '\xa0' with space in anchor tags content then appends in the degree names list.
		    except TypeError:
			degree_names.append(temp_name.contents[0])        
			continue
		degree_names = degree_names[23:]
"""
This function crawl the job title names and save in the list from the given url using Python BeautifulSoup library.
"""
	def crawl_job_title(self):		
		url = 'https://www.cebglobal.com/talent-management/talent-neuron/resources/job-posted-online.html' #Url for html content parsing.
		page = requests.get(url)
		soup = BeautifulSoup(page.text, 'html.parser') #It parses the html data from requested url.
		job__html_data = soup.find("section", {"id": "column_Section1"}) #It parses html data of specific class from all web page data.
		job__html_data.append(soup.find("section", {"id": "column_Section2"})) #It parses html data of specific class from all web page data.
		job__html_data = job_titles.find_all('li') #It parses html data of list tag from all html data with specific class.
		job_title_names = []
		for temp_name in job__html_data:
		    job_title_names.append(temp_name.contents[0]) #It appends anchor tag content in the job title names list.
"""
This function crawl the industries names and save in the list from the given url using Python BeautifulSoup library.
"""
	def craw_industries_name(self):		
		url = 'http://www.indianmirror.com/indian-industries/industries.html'
		page = requests.get(url)
		soup = BeautifulSoup(page.text, 'html.parser')
		industries_html_data = soup.find("table")
		industries_html_data = industries_html_data.find_all("a")
		industries_names = []
		for TempName in industries__html_data:
		    industries_names.append(TempName.contents[0])
"""
This function crawl the skills names and save in the list from the given url using Python BeautifulSoup library.
"""
	def craw_skills_name(self):	
		skills_names = []
		def parse_web_page_data(url):
		    page = requests.get(url)
		    soup = BeautifulSoup(page.text, 'html.parser') #It parses the html data from requested url.
		    skills_html_data = soup.find("div", {"class": "comp flex article-content expert-content"}) #It parses html data of specific class from all web page data.
		    skills_html_data = skills_html_data.find_all("li") #It parses html data of list tag from all html data with specific class.
		    for temp_name in skills_html_data:
			    anchor_tag_link = temp_name.find('a')
			    if anchor_tag_link == None:
				skills_names.append(temp_name.contents[0]) #It appends anchor tag content in the job title names list.
			    else:
				skills_names.append(anchor_tag_link.contents[0])    

		parse_web_page_data("https://www.thebalance.com/employability-skills-list-and-examples-4143571") #Url for html content parsing.
		parse_web_page_data("https://www.thebalance.com/human-resources-skills-list-2063754") #Url for html content parsing.
		parse_web_page_data("https://www.thebalance.com/list-of-general-skills-2063753") #Url for html content parsing.
		parse_web_page_data("https://www.thebalance.com/employment-skills-listed-by-job-2062389") #Url for html content parsing.
		skills_names = list(set(skills_names)) #Remove duplicancy in skills_names list.
"""
This function save the all crawled names in the candidate.xlsx file.
"""
	def save_crawls_fields_into_file(self):
		book = openpyxl.load_workbook('candidate.xlsx')
		sheet = book.active
		index = 2
		degree_size = len(degree_names) - 1 #Length - 1 of degree name list.
		institude_size = len(insitude_names) - 1 #Length - 1 of institude name list.
		job_title_size = len(job_title_names) - 1 #Length - 1 of job title name list.
		industries_size = len(industries_names) - 1 #Length - 1 of industries name list.
		skills_size = len(skills_names) - 1 #Length -1 of skills name list.
		max_experience = 20
		while index < sheet.max_col:
		    random_value = random.randint(0, job_title_size) 
		    sheet['C' + str(index)] = job_title_names[random_value] #Save random job title names in the excel file in 'C' field.
		    random_value = random.randint(0, industries_size)
		    sheet['D' + str(index)] = industries_names[random_value] #Save random industries title names in the excel file in 'D' field.
		    random_value = random.randint(0, max_experience)
		    sheet['E' + str(index)] = RandomValue * 12 #Save experience in months in the excel file in 'E' field.
		    random_value = random.randint(0, institude_size)
		    sheet['F' + str(index)] = institude_names[random_value] #Save random institude names in the excel file in 'F' field 	    
		    random_value = random.randint(0, degree_size)
		    sheet['G' + str(index)] = degree_names[random_value] #Save random degree names in the excel file in 'G' field.
		    random_skill_value = random.randint(2, 5)
		    skill_index_value = 0
		    skill_field_value = []
		    while skill_index_value <= random_skill_value:
			RandomValue = random.randint(0, SkillsMaxValue) 
			skill_field_value.append(skills_names[random_value])        
			skill_index_value += 1
		    skill_field_value = (", ".join(skill_field_value))
		    sheet['B' + str(index)] = skill_field_value    #Save random skills names in the excel file in 'B' field.
		    index += 1       
		book.save('candidate.xlsx') #Save all changes in the candidate.xlsx file.
"""
This function read job.xlsx file experince field which is in year then convert it in to max and min experience field in months and save in
two lists named min_exp_in_month and max_exp_in_month.
"""	
	def covert_experience_field_yr_to_month(self):
		book = openpyxl.load_workbook('job.xlsx') #Read job.xlsx file rows using load_workbook function of openpyxl module.	
		sheet = book.active #It sets current active sheet in job.xlsx file.
		max_row = sheet.max_row #It sets maximum number of rows in the job.xlsx file.
		index_value = 2
		min_exp_in_month = []
		max_exp_in_month = []
		error_cell_index = []

		#Loop which read all experience column cell value.
		while index_value < = max_row:
		    try:
			cell_value = sheet.cell(row=index_value, column=4)        
			cell_value = cell_value.value.split()
			cell_value = cell_value[:-1]  
			min_exp_in_month.append(int(cell_value[0]) * 12) #Convert cell string value into int and multiply with 12 and append to a list.
			max_exp_in_month.append(int(cell_value[2]) * 12) #Convert cell string value into int and multiply with 12 and append to a list.
			index_value += 1
		    except ValueError:
			error_cell_index.append(index_value) #Store unstructure cell index in error_cell_index variable
			index_value += 1
			continue
		    except IndexError:
			error_cell_index.append(index_value) #Store unstructure cell index in error_cell_index variable
			index_value += 1
			continue    
		    except AttributeError:
			error_cell_index.append(index_value) #Store unstructure cell index in error_cell_index variable
			index_value += 1
			continue
"""
This function read min_exp_in_month and max_exp_in_month which is in months then save in the job.xlsx file at F and G column.
"""
	def save_max_min_experinence_in_file(self):		
		book = openpyxl.load_workbook('job.xlsx')
		sheet = book.active
		index_value = 2
		list_index_value = 0
		max_row = sheet.max_row #It sets maximum number of rows in the job.xlsx file.
		while index_value <= max_row:
		    if index_value not in error_cell_index:			
			sheet['F' + str(index_value)] = min_exp_in_month[list_index_value] #It saves the min and max experience in F column of excel file
			sheet['G' + str(index_value)] = max_exp_in_month[list_index_value] #It saves the min and max experience in G column of excel file
			list_index_value += 1
		    index_value += 1   	    
		book.save("job.xlsx") #Save changes in excel file.
"""
This function unstructured data from the job.xlsx.
"""
	def space_at_unstructure_cell(self):
		book = openpyxl.load_workbook('job.xlsx')
		sheet = book.active
		index_value = 1
		column_letters = string.ascii_uppercase[:26] #Generate uppercase A-Z letters.
		column_letters = list(column_letters)
		char_index = 0
		while index < len(errorcellindex):        
		    while charindex < len(column_letters):
			sheet[column_letters[char_index] + str(errorcellindex[index_value])] = ""    #It gives blank at every unstructure data.
			char_index += 1
		    index_value += 1       
		    char_index = 0
		book.save("job.xlsx") #Save changes in excel file.

"""
This function split the dataset in 70:30 ratio and save in train and test.xlsx files.
"""
	def slipt_dataset_70_30(self, file_name):
		dataset = pd.read_excel(file_name) #Read excel file using pandas library object.
		file_name = file_name.split('.')
		train, test = train_test_split(dataset, test_size = 0.3) #Slipt dataset in 70:30 ratio and save in traina and test variable.
		train.to_csv(file_name + "_train.xlsx", index=False, index_label=False, header=True) #Save train dataset in the train.xlsx.
		test.to_csv(file_name + "_test.xlsx", index=False, index_label=False, header=True) #Save test dataset in the test.xlsx.
"""
This function convert the xlsx files into csv files as our requirement.
"""
	def convert_xlsx_to_csv(file_name):		
		wb = xlrd.open_workbook(file_name) #read excel file using xlrd open function.
		sh = wb.sheet_by_name('Sheet1') #load active sheet data.
		file_name = file_name.split('.') #split the file name.
		your_csv_file = open(file_name[0] + '.csv', 'w') #open csv file in write mode.
		wr = csv.writer(your_csv_file, delimiter="|") #give delimiter in excel file.

		for rownum in range(sh.nrows):
			wr.writerow(sh.row_values(rownum)) #convert every row of excel file and write into csv file.

		your_csv_file.close() #save all changes in csv file
"""
This function implement the collaborative filtering algorithm on the train and test data and recommend the jobs for the perfect candidate.
"""
	def collaborative_filtering_algo():		
		#Reading users file:
		u_cols = ['CV_Id', 'Skills', 'Designation', 'Industry', 'Experience (in months)', 'Institute', 'Degree']
		users = pd.read_csv('CandidateTrain.csv', sep='|', names=u_cols,
		 encoding='latin-1')

		#Reading job_mapping file:
		r_cols = ['JobId', 'CandidateID']
		job_mapping_train = pd.read_csv('job_mapping_train.csv', sep='|', names=r_cols, encoding='latin-1') #read train dataset. 
		job_mapping_test = pd.read_csv('job_mapping_test.csv', sep='|', names=r_cols, encoding='latin-1') #read test dataset.

		graphlab.product_key.set_product_key('EBFD-E604-0274-7909-2769-C6C5-0D2F-7516') # Set Graphlab product key
		train_data = graphlab.SFrame(job_mapping_train) #Load training data
		test_data = graphlab.SFrame(job_mapping_test) #Load testing data

		#Train Model
		item_sim_model = graphlab.item_similarity_recommender.create(train_data, user_id='JobId', item_id= "CandidateID", 	similarity_type='pearson')

		#Make Recommendations of first five users.
		item_sim_recomm = item_sim_model.recommend(users=range(1,6),k=5)
		item_sim_recomm.print_rows(num_rows=25)	

		#Measaure the performance of trained model
		model_performance = graphlab.compare(test_data, [item_sim_model])
	
